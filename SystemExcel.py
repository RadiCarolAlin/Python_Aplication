from openpyxl import load_workbook
from FunctionForArhitectExcel import *
import binascii
from GUI_Ppar import *
import global_


def parameterVerificationSystem(list):
    wx.Yield()  # GUI Refresh                                                           #to be written in the Output excel
    file = open("LogWarning.txt", "a")
    file.write("\n[SYSTEM SECTION]\n")
    for self in list:
        self.count = conversionCountFor2Value(self.type, self.count)
        auxType = conversionAndVerifyType(self.type)
        if auxType is False:
            file.write("[TYPE ERROR]    Name of the parameter: " + self.name + "   Type of the parameter: " + self.type + "\n")
        else:
            self.type = auxType

        if valid_data(self.type, self.count, self.value) is False:
            file.write("[OUT OF RANGE]  Name: " + str(self.name) + ", Type: " + str(self.type) + ", Count: " + str(self.count) + ", Value: " + str(self.value) + "\n")

        if isinstance(self.value, str) is True and ("0x" in self.value or "0X" in self.value) and ("CalSet".upper() not in self.name.upper() and "AntCompValue".upper() not in self.name.upper()):
            lengthValues(self.name, self.value, self.count, self.type, file)
    file.close()


def parameterVerificationOutput(listArchitect, listSystem):
    for system in listSystem:
        for arch in listArchitect:
            if system.name.upper() == arch.name.upper():
                arch.value = system.value

    wx.Yield()  # GUI Refresh                                                           #to be written in the Output excel
    file = open("LogWarning.txt", "a")
    file.write("\n[OUTPUT SECTION]\n")
    for self in listArchitect:
        auxType = conversionAndVerifyType(self.type)
        if auxType is False:
            file.write("[TYPE ERROR] " + "Name of the parameter: " + self.name + "  Type of the parameter: " + self.type + "\n")
            global_.stateOfTheProgram = "TYPE ERROR\n" + "Name of the parameter: " + self.name + "  Type of the parameter: " + self.type
            raise RuntimeError

        self.type = auxType

        if valid_data(self.type, self.count, self.value) is False:
            file.write("[OUT OF RANGE]  Name: " + str(self.name) + ", Type: " + str(self.type) + ", Count: " + str(self.count) + ", Value: " + str(self.value) + "\n")

        if isinstance(self.value, str) is True and ("0x" in self.value or "0X" in self.value) and ("CalSet".upper() not in self.name.upper() and "AntCompValue".upper() not in self.name.upper()):
            lengthValues(self.name, self.value, self.count, self.type, file)
    file.close()


class DataFromXlsSystem2:  # this class is created to read the System excel, store the elements in the class's
    def __init__(self, name, type, count, limit, value, TreeLevel2, lowLimit, maxLimit):  # members and write them in the Output excel
        self.name = name
        self.type = type
        self.count = count
        self.limit = limit
        self.value = value
        self.TreeLevel2 = TreeLevel2
        self.lowLimit = lowLimit
        self.maxLimit = maxLimit

    def __repr__(self):
        return "(" + str(self.name) + "," + str(self.type) + "," + str(self.count) + "," + str(self.limit) + "," + str(
            self.value) + "," + str(self.TreeLevel2) + "," + str(self.lowLimit) + "," + str(self.maxLimit) + ")"

    def readXMLSystem2(self):  # this function reads the parameters and stores them into variables
        listDataFromSystem2 = []
        listWithNotUsedElements2 = functionForListWithNotUsedElements(global_.listWithNotUsedElements)

        while global_.row2:
            wx.Yield()  # GUI Refresh
            self.value = ""
            self.name = global_.sheetSystem.cell(global_.row2, global_.nameColumn2).value
            self.type = global_.sheetSystem.cell(global_.row2, global_.typeColumn2).value
            self.count = global_.sheetSystem.cell(global_.row2, global_.countColumn2).value

            if self.count is not None:
                if isinstance(self.count, str):
                    if "=" in self.count:
                        self.count = eval(self.count[1:])

            intermediaryLimitRow = global_.row2 + 3
            self.limit = global_.sheetSystem.cell(intermediaryLimitRow, global_.limitValueColumn).value
            elementFound = 0

            if self.name is not None:
                rowLowAndMaxLimit = global_.row2

                self.lowLimit = global_.sheetSystem.cell(rowLowAndMaxLimit + 1, global_.valueColumn2).value
                if global_.sheetSystem.cell(rowLowAndMaxLimit + 2, global_.limitValueColumn).value == "upper":   # pragma: no cover
                    self.maxLimit = global_.sheetSystem.cell(rowLowAndMaxLimit + 2, global_.valueColumn2).value
                else:
                    rowLowAndMaxLimit += 3
                    while global_.sheetSystem.cell(rowLowAndMaxLimit, global_.limitValueColumn).value != "upper":
                        rowLowAndMaxLimit += 1
                    self.maxLimit = global_.sheetSystem.cell(rowLowAndMaxLimit, global_.valueColumn2).value

                if isinstance(self.lowLimit, str):
                    if self.lowLimit is not None:
                        if "," in self.lowLimit:
                            index = self.lowLimit.index(",")
                            self.lowLimit = self.lowLimit[:index]
                    self.lowLimit = twosComplement_hex(self.lowLimit)

                if isinstance(self.maxLimit, str):
                    if self.maxLimit is not None:
                        if "," in self.maxLimit:
                            index = self.maxLimit.index(",")
                            self.maxLimit = self.maxLimit[:index]
                    self.maxLimit = twosComplement_hex(self.maxLimit)

            if self.name is not None:
                if str(self.limit) != "upper" and self.limit is not None:
                    while self.limit != "upper":
                        if global_.projectSpecificName.strip().upper() == self.limit.strip().upper() or global_.projectNameType.strip().upper() == self.limit.strip().upper():
                            elementFound = 1
                            self.value = self.value + str(global_.sheetSystem.cell(row=intermediaryLimitRow, column=global_.valueColumn2).value)
                            if global_.sheetSystem.cell(intermediaryLimitRow, global_.valueColumn2).value is not None:
                                if (global_.sheetSystem.cell(intermediaryLimitRow, global_.valueColumn2 + 1)) is not None:
                                    valueColumnIntermediary2 = global_.valueColumn2
                                    value2 = ""
                                    while value2 is not None:
                                        valueColumnIntermediary2 = valueColumnIntermediary2 + 1
                                        value2 = global_.sheetSystem.cell(intermediaryLimitRow, valueColumnIntermediary2).value
                                        if value2 is not None:   # pragma: no cover
                                            self.value = self.value + ";" + str(value2)

                        elif "\n" in self.limit:
                            listWithLimits = re.split(' |/|\n', str(self.limit))
                            for element in listWithLimits:
                                if element.strip().upper() == global_.projectSpecificName.strip().upper():
                                    elementFound = 1
                                    self.value = self.value + str(global_.sheetSystem.cell(row=intermediaryLimitRow, column=global_.valueColumn2).value)
                                    if global_.sheetSystem.cell(intermediaryLimitRow, global_.valueColumn2).value is not None:
                                        if (global_.sheetSystem.cell(intermediaryLimitRow, global_.valueColumn2 + 1)) is not None:
                                            valueColumnIntermediary2 = global_.valueColumn2
                                            value2 = ""
                                            while value2 is not None:
                                                valueColumnIntermediary2 = valueColumnIntermediary2 + 1
                                                value2 = global_.sheetSystem.cell(intermediaryLimitRow,
                                                                                  valueColumnIntermediary2).value
                                                if value2 is not None:   # pragma: no cover
                                                    self.value = self.value + ";" + str(value2)
                        intermediaryLimitRow += 1
                        self.limit = global_.sheetSystem.cell(intermediaryLimitRow, global_.limitValueColumn).value

                if str(self.limit) == "upper" and elementFound == 0:
                    self.value = self.value + str(global_.sheetSystem.cell(row=global_.row2 + 2, column=global_.valueColumn2).value)
                    if global_.sheetSystem.cell(global_.row2 + 2, global_.valueColumn2).value is not None:
                        if (global_.sheetSystem.cell(global_.row2 + 2, global_.valueColumn2 + 1)) is not None:
                            valueColumnIntermediary = global_.valueColumn2
                            value1 = ""
                            while value1 is not None:
                                valueColumnIntermediary = valueColumnIntermediary + 1
                                value1 = global_.sheetSystem.cell(global_.row2 + 2, valueColumnIntermediary).value
                                if value1 is not None:
                                    self.value = self.value + ";" + str(value1)
                                else:
                                    if self.name == "Monopulse Phase Azimuth":
                                        self.lowLimit = global_.sheetSystem.cell(global_.row2 + 1, valueColumnIntermediary - 1).value
                                        self.lowLimit = twosComplement_hex(self.lowLimit)

            if self.type == "UI64" or self.type == "SI64":
                self.value = conversionInt64(self.value)

            if self.name is not None and self.type is None and self.count is None and self.limit is None:
                self.TreeLevel2 = str(self.name)

            if self.name is not None:
                if "not to be used by SW" in self.name:
                    index = self.name.index("not to be used by SW")
                    self.name = self.name[:index].strip()

            functionForListWithNotUsedElements(global_.listWithNotUsedElements)
            if self.name is not None:
                if self.name.strip() not in listWithNotUsedElements2:
                    if self.type is not None:
                        if self.value is not None and self.value != "":
                            if "None" not in self.value:

                                if self.name == "HM_HomologationMonitor_CRC":
                                    aux = self.value[2:].upper()
                                    self.value = "0x" + aux

                                Object = DataFromXlsSystem2(self.name, self.type, self.count, self.limit, self.value,
                                                            self.TreeLevel2, self.lowLimit, self.maxLimit)
                                listDataFromSystem2.append(Object)

            global_.row2 = global_.row2 + 1
            if self.name == global_.rowEnd2:
                global_.row2 = False
        return listDataFromSystem2

    @staticmethod
    def writeInExcel2(listDataFromSystem2):  # this function writes the data retrieved from the System excel in the Output excel
        global_.sheetOutput = global_.excelWriteOutput[global_.output1Sheet]
        global_.sheetOutput2 = global_.excelWriteOutput[global_.output2Sheet]

        parameterVerificationSystem(listDataFromSystem2)

        for self in listDataFromSystem2:
            wx.Yield()
            rowWrite2 = 1
            while rowWrite2:
                nameFromExcel = global_.sheetOutput.cell(rowWrite2, global_.nameColumnWrite).value
                valueTreeLevel2 = global_.sheetOutput.cell(rowWrite2, global_.writeTreeLevel2).value
                if nameFromExcel is not None:
                    if valueTreeLevel2 == self.TreeLevel2:
                        if self.name in nameFromExcel and len(self.name.strip()) == len(nameFromExcel.strip()):
                            if self.count is not None:

                                if self.lowLimit is not None and self.maxLimit is not None:
                                    global_.sheetOutput.cell(row=rowWrite2, column=global_.lowLimitColumn, value=self.lowLimit)
                                    global_.sheetOutput.cell(row=rowWrite2, column=global_.maxLimitColumn, value=self.maxLimit)

                                if self.count == 1:
                                    if isinstance(self.value, str) is True and "0x" not in self.value:
                                        if self.value[0] == "=":
                                            self.value = eval(self.value[1:])
                                        else:
                                            self.value = eval(self.value)
                                    global_.sheetOutput.cell(row=rowWrite2, column=global_.valueColumnWrite, value=self.value)

                                    break
                                elif 32 >= self.count > 1:
                                    if self.value != "None":
                                        listWithValue = re.split(';|,', str(self.value))
                                        columnWriteIterator = global_.valueColumnWrite
                                        for value in listWithValue:
                                            if "0x" not in value:
                                                value = int(value)
                                            global_.sheetOutput.cell(row=rowWrite2, column=columnWriteIterator, value=value)
                                            columnWriteIterator = columnWriteIterator + 4
                                        break
                                elif self.count > 32:
                                    if self.value != "None":
                                        column = 2
                                        if global_.sheetOutput.cell(rowWrite2, global_.valueColumnWrite).value == "add_PPAR_default_data":
                                            while global_.sheetOutput2.cell(2, column).value is not None:
                                                if global_.sheetOutput2.cell(2, column).value.upper().strip() == self.name.upper().strip():
                                                    listWithValue = re.split(';|,', str(self.value))
                                                    intermediaryRowWriteSheet2 = global_.rowWriteSheet2 + 1
                                                    contor = 0
                                                    for valueAux in listWithValue:   # pragma: no cover
                                                        if self.type == "SI16":
                                                            if valueAux[0:2] == "0x" and valueAux != "0x00":
                                                                valueAux = valueAux[2:6]
                                                                valueAux = bytes(valueAux, 'utf-8')
                                                                valueAux = binascii.a2b_hex(valueAux)
                                                                valueAux = (int.from_bytes(valueAux, byteorder='big', signed=True))
                                                                listWithValue[contor] = valueAux
                                                                contor += 1

                                                    for value in listWithValue:
                                                        if not isinstance(listWithValue[0], int):
                                                            if value[0:2] != "0x":
                                                                value = int(value)

                                                        global_.sheetOutput2.cell(row=intermediaryRowWriteSheet2, column=column,
                                                                                  value=value)
                                                        intermediaryRowWriteSheet2 = intermediaryRowWriteSheet2 + 1
                                                    break
                                                column += 1

                                        else:    # pragma: no cover
                                            global_.sheetOutput.cell(row=rowWrite2, column=global_.valueColumnWrite,
                                                                     value="add_PPAR_default_data")
                                            listWithValue = re.split(';|,', str(self.value))

                                            global_.sheetOutput2.cell(row=1, column=global_.columnWriteSheet2, value=self.TreeLevel2)
                                            global_.sheetOutput2.cell(row=global_.rowWriteSheet2, column=global_.columnWriteSheet2, value=self.name)

                                            intermediaryRowWriteSheet2 = global_.rowWriteSheet2 + 1
                                            contor = 0
                                            for valueAux in listWithValue:   # pragma: no cover
                                                if self.type == "SI16":
                                                    if valueAux[0:2] == "0x" and valueAux != "0x00":
                                                        valueAux = valueAux[2:6]
                                                        valueAux = bytes(valueAux, 'utf-8')
                                                        valueAux = binascii.a2b_hex(valueAux)
                                                        valueAux = (int.from_bytes(valueAux, byteorder='big', signed=True))
                                                        listWithValue[contor] = valueAux
                                                        contor += 1

                                            for value in listWithValue:
                                                if not isinstance(listWithValue[0], int):
                                                    if value[0:2] != "0x":
                                                        value = int(value)

                                                global_.sheetOutput2.cell(row=intermediaryRowWriteSheet2, column=global_.columnWriteSheet2,
                                                                          value=value)
                                                intermediaryRowWriteSheet2 = intermediaryRowWriteSheet2 + 1
                                            global_.columnWriteSheet2 = global_.columnWriteSheet2 + 1
                                    else:    # pragma: no cover
                                        break

                rowWrite2 = rowWrite2 + 1
                if rowWrite2 > 800:
                    break
        global_.excelWriteOutput.save(global_.pathOutput)
        return global_.columnWriteSheet2
