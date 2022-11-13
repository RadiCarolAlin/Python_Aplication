import unittest
from openpyxl import load_workbook
import ArchitectExcel
import FunctionForArhitectExcel
import SystemExcel
from ArchitectExcel import DataFromXlsArchitect
import OemSpecific
import global_
import GUI_Ppar
import A2LColumn

class Testmain(unittest.TestCase):

    def test_reverse_001(self):
        result = FunctionForArhitectExcel.reverse("abcd")
        self.assertEqual(result, "dcba")

    def test_reverse_002(self):
        result = FunctionForArhitectExcel.reverse(" ABCD")
        self.assertEqual(result, "DCBA ")

    def test_hexCoresponded_001(self):
        result = FunctionForArhitectExcel.hexCorrespond(10)
        self.assertEqual(result, "A")

    def test_hexCoresponded_002(self):
        result = FunctionForArhitectExcel.hexCorrespond(11)
        self.assertEqual(result, "B")

    def test_hexCoresponded_003(self):
        result = FunctionForArhitectExcel.hexCorrespond(12)
        self.assertEqual(result, "C")

    def test_hexCoresponded_004(self):
        result = FunctionForArhitectExcel.hexCorrespond(13)
        self.assertEqual(result, "D")

    def test_hexCoresponded_005(self):
        result = FunctionForArhitectExcel.hexCorrespond(14)
        self.assertEqual(result, "E")

    def test_hexCoresponded_006(self):
        result = FunctionForArhitectExcel.hexCorrespond(15)
        self.assertEqual(result, "F")

    def test_hexCoresponded_007(self):
        result = FunctionForArhitectExcel.hexCorrespond(9)
        self.assertEqual(result, 9)

    def test_conversionType_001(self):
        result = FunctionForArhitectExcel.conversionAndVerifyType("UI8")
        self.assertEqual(result, "uint8")

    def test_conversionType_002(self):
        result = FunctionForArhitectExcel.conversionAndVerifyType("SI8")
        self.assertEqual(result, "sint8")

    def test_conversionType_003(self):
        result = FunctionForArhitectExcel.conversionAndVerifyType("UI16")
        self.assertEqual(result, "uint16")

    def test_conversionType_004(self):
        result = FunctionForArhitectExcel.conversionAndVerifyType("SI16")
        self.assertEqual(result, "sint16")

    def test_conversionType_005(self):
        result = FunctionForArhitectExcel.conversionAndVerifyType("UI32")
        self.assertEqual(result, "uint32")

    def test_conversionType_006(self):
        result = FunctionForArhitectExcel.conversionAndVerifyType("SI32")
        self.assertEqual(result, "sint32")

    def test_conversionType_007(self):
        result = FunctionForArhitectExcel.conversionAndVerifyType("FL32")
        self.assertEqual(result, "float32")

    def test_conversionType_008(self):
        result = FunctionForArhitectExcel.conversionAndVerifyType("UI64")
        self.assertEqual(result, "uint64")

    def test_conversionType_009(self):
        result = FunctionForArhitectExcel.conversionAndVerifyType("SI64")
        self.assertEqual(result, "sint64")

    def test_conversionType_010(self):
        result = FunctionForArhitectExcel.conversionAndVerifyType("2x SI16")
        self.assertEqual(result, "sint16")

    def test_conversionType_011(self):
        result = FunctionForArhitectExcel.conversionAndVerifyType("testType")
        self.assertEqual(result, False)

    def test_completeNoneValue_001(self):
        result = FunctionForArhitectExcel.completeNoneValue("uint8", None, "0xFF")
        self.assertEqual(result, "0xFF")

    def test_completeNoneValue_002(self):
        result = FunctionForArhitectExcel.completeNoneValue("sint8", None, "0xFF")
        self.assertEqual(result, "0xFF")

    def test_completeNoneValue_003(self):
        result = FunctionForArhitectExcel.completeNoneValue("uint16", None, "0xFF")
        self.assertEqual(result, "0xFFFF")

    def test_completeNoneValue_004(self):
        result = FunctionForArhitectExcel.completeNoneValue("sint16", None, "0xFF")
        self.assertEqual(result, "0xFFFF")

    def test_completeNoneValue_005(self):
        result = FunctionForArhitectExcel.completeNoneValue("uint32", None, "0xFF")
        self.assertEqual(result, "0xFFFFFFFF")

    def test_completeNoneValue_006(self):
        result = FunctionForArhitectExcel.completeNoneValue("sint32", None, "0xFF")
        self.assertEqual(result, "0xFFFFFFFF")

    def test_completeNoneValue_007(self):
        result = FunctionForArhitectExcel.completeNoneValue("float32", None, "0xFF")
        self.assertEqual(result, "0xFFFFFFFF")

    def test_completeNoneValue_008(self):
        result = FunctionForArhitectExcel.completeNoneValue("sint64", None, "0xFF")
        self.assertEqual(result, "0xFFFFFFFFFFFFFFFF")

    def test_completeNoneValue_009(self):
        result = FunctionForArhitectExcel.completeNoneValue("uint64", None, "0xFF")
        self.assertEqual(result, "0xFFFFFFFFFFFFFFFF")

    def test_completeNoneValue_010(self):
        result = FunctionForArhitectExcel.completeNoneValue("uint8", None, "0x00")
        self.assertEqual(result, "0x00")

    def test_completeNoneValue_011(self):
        result = FunctionForArhitectExcel.completeNoneValue("uint8", "0x0", "0x00")
        self.assertEqual(result, "0x00")

    def test_completeNoneValue_012(self):
        result = FunctionForArhitectExcel.completeNoneValue("sint8", "0x0", "0xFF")
        self.assertEqual(result, "0x00")

    def test_completeNoneValue_013(self):
        result = FunctionForArhitectExcel.completeNoneValue("uint16", "0x0", "0xFF")
        self.assertEqual(result, "0x0000")

    def test_completeNoneValue_014(self):
        result = FunctionForArhitectExcel.completeNoneValue("sint16", "0x0", "0xFFFF")
        self.assertEqual(result, "0x0000")

    def test_completeNoneValue_015(self):
        result = FunctionForArhitectExcel.completeNoneValue("uint32", "0x0", "0xFF")
        self.assertEqual(result, "0x00000000")

    def test_completeNoneValue_016(self):
        result = FunctionForArhitectExcel.completeNoneValue("sint32", "0x0", "0xFF")
        self.assertEqual(result, "0x00000000")

    def test_completeNoneValue_017(self):
        result = FunctionForArhitectExcel.completeNoneValue("float32", "0x0", None)
        self.assertEqual(result, "0x00000000")

    def test_completeNoneValue_018(self):
        result = FunctionForArhitectExcel.completeNoneValue("sint64", "0x0", "0xFF")
        self.assertEqual(result, "0x0000000000000000")

    def test_completeNoneValue_019(self):
        result = FunctionForArhitectExcel.completeNoneValue("uint64", "PROD", "0xFF")
        self.assertEqual(result, "0x0000000000000000")

    def test_completeNoneValue_020(self):
        result = FunctionForArhitectExcel.completeNoneValue("uint8", "NA", "0xFF")
        self.assertEqual(result, "0xFF")

    def test_completeNoneValue_021(self):
        result = FunctionForArhitectExcel.completeNoneValue("sint8", "NA", "0xFF")
        self.assertEqual(result, "0xFF")

    def test_completeNoneValue_022(self):
        result = FunctionForArhitectExcel.completeNoneValue("uint16", "NA", "0xFF")
        self.assertEqual(result, "0xFFFF")

    def test_completeNoneValue_023(self):
        result = FunctionForArhitectExcel.completeNoneValue("sint16", "NA", "0xFFFF")
        self.assertEqual(result, "0xFFFF")

    def test_completeNoneValue_024(self):
        result = FunctionForArhitectExcel.completeNoneValue("uint32", "NA", "0xFF")
        self.assertEqual(result, "0xFFFFFFFF")

    def test_completeNoneValue_025(self):
        result = FunctionForArhitectExcel.completeNoneValue("sint32", "NA", "0xFF")
        self.assertEqual(result, "0xFFFFFFFF")

    def test_completeNoneValue_026(self):
        result = FunctionForArhitectExcel.completeNoneValue("float32", "NA", "0xFF")
        self.assertEqual(result, "0xFFFFFFFF")

    def test_completeNoneValue_027(self):
        result = FunctionForArhitectExcel.completeNoneValue("sint64", "N/A\\0", "0xFF")
        self.assertEqual(result, "0xFFFFFFFFFFFFFFFF")

    def test_completeNoneValue_028(self):
        result = FunctionForArhitectExcel.completeNoneValue("uint64", "N/A\\0", None)
        self.assertEqual(result, "0xFFFFFFFFFFFFFFFF")

    def test_conversionValue_001(self):
        result = FunctionForArhitectExcel.conversionValue("uint8", 1, "0xFF")
        self.assertEqual(result, "0xFF")

    def test_conversionValue_002(self):
        result = FunctionForArhitectExcel.conversionValue("uint16", 1, "0xAB")
        self.assertEqual(result, "0x00AB")

    def test_conversionValue_003(self):
        result = FunctionForArhitectExcel.conversionValue("uint32", 1, "0xCC")
        self.assertEqual(result, "0x000000CC")

    def test_conversionValue_004(self):
        result = FunctionForArhitectExcel.conversionValue("uint32", 1, "test")
        self.assertEqual(result, "0x74657374")

    def test_conversionValue_005(self):
        result = FunctionForArhitectExcel.conversionValue("x", 4, "\"ARS510\\0")
        self.assertEqual(result, "0x4152533531305C30")

    def test_conversionValue_006(self):
        result = FunctionForArhitectExcel.conversionValue("x", 4, "0XBB")
        self.assertEqual(result, "0xBB")

    def test_conversionValue_007(self):
        result = FunctionForArhitectExcel.conversionValue("x", 4, "AB12")
        self.assertEqual(result, "0x41423132")

    def test_conversionValue_008(self):
        result = FunctionForArhitectExcel.conversionValue("uint8", 2, "0xFF")
        self.assertEqual(result, "0xFFFF")

    def test_conversionValue_009(self):
        result = FunctionForArhitectExcel.conversionValue("uint16", 2, "0xAB")
        self.assertEqual(result, "0x00AB00AB")

    def test_conversionValue_010(self):
        result = FunctionForArhitectExcel.conversionValue("uint32", 2, "0xCCCC")
        self.assertEqual(result, "0x0000CCCC0000CCCC")

    def test_conversionInt32Adress_001(self):
        result = FunctionForArhitectExcel.conversionInt32Address(123)
        self.assertEqual(result, "0x7B")

    def test_conversionAsciiToHex_001(self):
        result = FunctionForArhitectExcel.conversionAsciiToHex(1, "\"0xF\\F")
        self.assertEqual(result, "0x307846")

    def test_conversionAsciiToHex_002(self):
        result = FunctionForArhitectExcel.conversionAsciiToHex(1, "test")
        self.assertEqual(result, "0x74657374")

    def test_conversionAsciiToHex_003(self):
        result = FunctionForArhitectExcel.conversionAsciiToHex(1, "0xF\\F")
        self.assertEqual(result, "0x307846")

    def test_conversionAsciiToHexSecondMethod_001(self):
        result = FunctionForArhitectExcel.conversionAsciiToHexSecondMethod(1, "0xF\"F")
        self.assertEqual(result, "0x7846")

    def test_conversionCountFor2Value_001(self):
        result = FunctionForArhitectExcel.conversionCountFor2Value("2x SI16", 12)
        self.assertEqual(result, 24)

    def test_multiplyHex_001(self):
        result = FunctionForArhitectExcel.multiplyHex("0xFF", 2)
        self.assertEqual(result, "0xFFFF")

    def test_conversionInt64_001(self):
        result = FunctionForArhitectExcel.conversionInt64(7.60E+10)
        self.assertEqual(result, "0x00000011B1F3F800")

    def test_conversionInt32_001(self):
        result = FunctionForArhitectExcel.conversionInt32("0x0000")
        self.assertEqual(result, "0x00000000")

    def test_conversionInt8_001(self):
        result = FunctionForArhitectExcel.conversionInt8("0xFF")
        self.assertEqual(result, "0xFF")

    def test_conversionInt16_001(self):
        result = FunctionForArhitectExcel.conversionInt16("0x00")
        self.assertEqual(result, "0x0000")

    def test_conversionOem2_001(self):
        result = FunctionForArhitectExcel.conversionOem2("\"002\" [ASCII]", 3)
        self.assertEqual(result, "0x303032")

    def test_conversionOem2_002(self):
        result = FunctionForArhitectExcel.conversionOem2("0xFF ", 16)
        self.assertEqual(result, "0xFF")

    def test_conversionOem_001(self):
        result = FunctionForArhitectExcel.conversionOem("\"0001\" [ASCII]", 4)
        self.assertEqual(result, "0x30303031")

    def test_conversionOem_002(self):
        result = FunctionForArhitectExcel.conversionOem("13.12.17[ASCII]", 4)
        self.assertEqual(result, "0x31332E31322E3137")

    def test_functionForListWithNotUsedElements_001(self):
        list1 = "el1,el2,el3"
        result = FunctionForArhitectExcel.functionForListWithNotUsedElements(list1)
        self.assertEqual(result, ["el1", "el2", "el3"])

    def test_twosComplement_hex_001(self):
        result = FunctionForArhitectExcel.twosComplement_hex("0xFFF6")
        self.assertEqual(result, -10)\

    def test_range_verification_001(self):
        result = FunctionForArhitectExcel.range_verification("uint8", 120)
        self.assertEqual(result, True)

    def test_range_verification_002(self):
        result = FunctionForArhitectExcel.range_verification("sint8", 127)
        self.assertEqual(result, True)

    def test_range_verification_003(self):
        result = FunctionForArhitectExcel.range_verification("uint16", 63500)
        self.assertEqual(result, True)

    def test_range_verification_004(self):
        result = FunctionForArhitectExcel.range_verification("sint16", -17000)
        self.assertEqual(result, True)

    def test_range_verification_005(self):
        result = FunctionForArhitectExcel.range_verification("uint32", 4294967)
        self.assertEqual(result, True)

    def test_range_verification_006(self):
        result = FunctionForArhitectExcel.range_verification("sint32", -1234567)
        self.assertEqual(result, True)

    def test_range_verification_007(self):
        result = FunctionForArhitectExcel.range_verification("float32", 320.0)
        self.assertEqual(result, True)

    def test_range_verification_008(self):
        result = FunctionForArhitectExcel.range_verification("uint64", 1)
        self.assertEqual(result, True)

    def test_range_verification_009(self):
        result = FunctionForArhitectExcel.range_verification("sint64", 922337203685)
        self.assertEqual(result, True)

    def test_range_verification_010(self):
        result = FunctionForArhitectExcel.range_verification("uint8", 1236859)
        self.assertEqual(result, False)

    def test_valid_data_001(self):
        result = FunctionForArhitectExcel.valid_data("uint8", 1, 1234)
        self.assertEqual(result, False)

    def test_valid_data_002(self):
        result = FunctionForArhitectExcel.valid_data("uint8", 1, "0x0118")
        self.assertEqual(result, False)

    def test_valid_data_003(self):
        result = FunctionForArhitectExcel.valid_data("sint8", 2, "0xA2B32") #pt verificare de lungime
        self.assertEqual(result, False)

    def test_valid_data_004(self):
        result = FunctionForArhitectExcel.valid_data("sint8", 2, "0x3A9A") #cazul fara ; cu o valaorea in range iar una in afara range-ului
        self.assertEqual(result, False)

    def test_valid_data_005(self):
        result = FunctionForArhitectExcel.valid_data("sint8", 2, "0xF2;0xA1C1") #caz cu ; cu o valoare corecta iar una nu
        self.assertEqual(result, False)

    def test_valid_data_006(self):
        result = FunctionForArhitectExcel.valid_data("sint8", 3, "0xF2E3;0xA1C1") #caz cu ; in care lungimea nu este egala cu count*2
        self.assertEqual(result, False)

    def test_valid_data_007(self):
        result = FunctionForArhitectExcel.valid_data("uint8", 2, 348)
        self.assertEqual(result, False)

    def test_valid_data_008(self):
        result = FunctionForArhitectExcel.valid_data("uint16", 1, "0XAD6F8")
        self.assertEqual(result, False)

    def test_valid_data_009(self):
        result = FunctionForArhitectExcel.valid_data("uint16", 1, 467897)
        self.assertEqual(result, False)

    def test_valid_data_010(self):
        result = FunctionForArhitectExcel.valid_data("sint8", 3, "0x12;FF")
        self.assertEqual(result, False)

    def test_valid_data_011(self):
        result = FunctionForArhitectExcel.valid_data("sint16", 2, "0X6F4DF7F9")
        self.assertEqual(result, False)

    def test_valid_data_012(self):
        result = FunctionForArhitectExcel.valid_data("sint16", 2, "0X6F4DF7F9;0X6F4DF7F9")
        self.assertEqual(result, False)

    def test_valid_data_013(self):
        result = FunctionForArhitectExcel.valid_data("sint16", 2, "0X6F4DF7F96")
        self.assertEqual(result, False)

    def test_valid_data_014(self):
        result = FunctionForArhitectExcel.valid_data("uint16", 2, 761456)
        self.assertEqual(result, False)

    def test_valid_data_015(self):
        result = FunctionForArhitectExcel.valid_data("uint32", 1, -26)
        self.assertEqual(result, False)

    def test_valid_data_016(self):
        result = FunctionForArhitectExcel.valid_data("sint32", 1, "0xF9F5EEFF")
        self.assertEqual(result, False)

    def test_valid_data_017(self):
        result = FunctionForArhitectExcel.valid_data("sint8", 3, "0x12;FFF")
        self.assertEqual(result, False)

    def test_valid_data_018(self):
        result = FunctionForArhitectExcel.valid_data("sint32", 2, "0x1A1111A6F9F5EEFF")
        self.assertEqual(result, False)

    def test_valid_data_019(self):
        result = FunctionForArhitectExcel.valid_data("sint32", 2, "0x1A1111A6F9F5EEFF;0x1A1111A6F9F5EEFF")
        self.assertEqual(result, False)

    def test_valid_data_020(self):
        result = FunctionForArhitectExcel.valid_data("sint32", 2, "0xFFFFFFFF;0x1A1111A6F9F5EEFF")
        self.assertEqual(result, False)

    def test_valid_data_021(self):
        result = FunctionForArhitectExcel.valid_data("uint32", 2, 56294967295)
        self.assertEqual(result, False)

    def test_valid_data_022(self):
        result = FunctionForArhitectExcel.valid_data("uint64", 1, -16579)
        self.assertEqual(result, False)

    def test_valid_data_023(self):
        result = FunctionForArhitectExcel.valid_data("uint64", 1, "0X1A1111A6F9F5EEFFA12")
        self.assertEqual(result, False)

    def test_valid_data_024(self):
        result = FunctionForArhitectExcel.valid_data("sint16", 2, "0XFFFF;0xFF")
        self.assertEqual(result, False)

    def test_valid_data_025(self):
        result = FunctionForArhitectExcel.valid_data("sint64", 2, "0X1A1A2B2C4D6E7F6C889A6D1A1A2B2C4D")
        self.assertEqual(result, False)

    def test_valid_data_026(self):
        result = FunctionForArhitectExcel.valid_data("sint64", 2, "0X1A1A2B2C4D6E7F6C889A6D1A1A2B2C4DD")
        self.assertEqual(result, False)

    def test_valid_data_027(self):
        result = FunctionForArhitectExcel.valid_data("sint16", 2, "0X1111;FFFF")
        self.assertEqual(result, False)

    def test_valid_data_028(self):
        result = FunctionForArhitectExcel.valid_data("sint64", 2, 922337203685477580767)
        self.assertEqual(result, False)

    def test_valid_data_029(self):
        result = FunctionForArhitectExcel.valid_data("float32", 1, 3.4028235 * pow(10, 38) * 100)
        self.assertEqual(result, False)

    def test_valid_data_030(self):
        result = FunctionForArhitectExcel.valid_data("uint8", 2, "800;567")
        self.assertEqual(result, False)

    def test_valid_data_031(self):
        result = FunctionForArhitectExcel.valid_data("float32", 1, "0x7F8FFFFF")
        self.assertEqual(result, False)

    def test_valid_data_032(self):
        result = FunctionForArhitectExcel.valid_data("float32", 2, str(0x10000000) + str( 3.4028235 * pow(10, 38) * 100))
        self.assertEqual(result, False)

    def test_valid_data_033(self):
        result = FunctionForArhitectExcel.valid_data("float32",2, "0x7F7FFFFF7F8FFFFF")
        self.assertEqual(result, False)

    def test_valid_data_034(self):
        result = FunctionForArhitectExcel.valid_data("uint16", 2, "6783245;223")
        self.assertEqual(result, False)

    def test_valid_data_035(self):
        result = FunctionForArhitectExcel.valid_data("uint32", 2, "783222383658923;2989352835")
        self.assertEqual(result, False)

    def test_valid_data_036(self):
        result = FunctionForArhitectExcel.valid_data("sint64", 2, "-143235643465789758697809;2741395")
        self.assertEqual(result, False)

    def test_valid_data_037(self):
        result = FunctionForArhitectExcel.valid_data("float32", 2, 3.4028235 * pow(10, 38) * 100)
        self.assertEqual(result, False)

    def test_valid_data_038(self):
        result = FunctionForArhitectExcel.valid_data("float32", 2, "0X00000000;0XFFFFFFFF")
        self.assertEqual(result, False)

    def test_valid_data_039(self):
        result = FunctionForArhitectExcel.valid_data("float32", 2, str(3.4028235 * pow(10, 38) * 100) + ";" + str(671349456789031039.4))
        self.assertEqual(result, False)

    def test_valid_data_040(self):
        result = FunctionForArhitectExcel.valid_data("sint16", 2, "0X1111;FFFFF")
        self.assertEqual(result, False)

    def test_valid_data_041(self):
        result = FunctionForArhitectExcel.valid_data("sint32", 2, "0xF9F5EEEFF")
        self.assertEqual(result, False)

    def test_valid_data_042(self):
        result = FunctionForArhitectExcel.valid_data("sint32", 2, "0x11111111;FFFFFFFF")
        self.assertEqual(result, False)

    def test_valid_data_043(self):
        result = FunctionForArhitectExcel.valid_data("sint32", 2, "0x11111111;FFFFFFFFF")
        self.assertEqual(result, False)

    def test_valid_data_044(self):
        result = FunctionForArhitectExcel.valid_data("sint64", 2, "0X1A1A2B2C4D6E7F6C889A6D1A1A2B2C4DD;0XFF")
        self.assertEqual(result, False)

    def test_valid_data_045(self):
        result = FunctionForArhitectExcel.valid_data("sint64", 2, "0XFFFFFFFFFFFFFFFF;0XFF")
        self.assertEqual(result, False)

    def test_valid_data_046(self):
        result = FunctionForArhitectExcel.valid_data("sint64", 2, "0X1111111111111111;FF")
        self.assertEqual(result, False)

    def test_valid_data_047(self):
        result = FunctionForArhitectExcel.valid_data("sint64", 2, "0X1111111111111111;FFFFFFFFFFFFFFFF")
        self.assertEqual(result, False)

    def test_valid_data_048(self):
        result = FunctionForArhitectExcel.valid_data("float32", 2, "0X00000000F")
        self.assertEqual(result, False)

    def test_valid_data_049(self):
        result = FunctionForArhitectExcel.valid_data("float32", 2, "0X00000000F;0XFF")
        self.assertEqual(result, False)

    def test_valid_data_050(self):
        result = FunctionForArhitectExcel.valid_data("float32", 2, "0X00000000;FF")
        self.assertEqual(result, False)

    def test_valid_data_051(self):
        result = FunctionForArhitectExcel.valid_data("float32", 2, "0X00000001;FFFFFFFF")
        self.assertEqual(result, False)

    def test_valid_data_052(self):
        result = FunctionForArhitectExcel.valid_data("uint8", 1, "0XFF")
        self.assertEqual(result, True)

    def test_takeFirstAdress_001(self):
        global_.excelWriteOutput = load_workbook(filename="UnitTest/xPAR_TEST.xlsx")
        global_.sheetOutput = global_.excelWriteOutput[global_.output1Sheet]
        global_.excelReadArchitect = load_workbook(filename="UnitTest/ArchitectTEST.xlsx",
                                                   data_only=True)
        global_.sheetArchitect = global_.excelReadArchitect[global_.architectSheet]
        global_.readRowAddress = 4
        global_.readColumnAddress = 10
        global_.writeRowAddress = 5
        global_.writeColumnAddress = 148
        ArchitectExcel.takeFirstAddress()

    def test_repr_001(self):
        self.Object = ArchitectExcel.DataFromXlsArchitect("TEST", "UI8", 1, "0x0", "Infoblock", "PPAR_ChannelVariationData", "0xFFFF")
        print(self.Object)

    def test_paramaterAdaptation_001(self):
        self.Object = ArchitectExcel.DataFromXlsArchitect("TEST", "dg8", 1, "0x0", "Infoblock", "PPAR_ChannelVariationData", "0xFFFF")
        testLista = []
        testLista.append(self.Object)
        with self.assertLogs(level='WARNING') as captured:
            #with self.assertRaises(RuntimeError): #2nd method to test if an error is raised.
                #ArchitectExcel.parameterAdaptation(testLista)
                self.assertRaises(RuntimeError, ArchitectExcel.parameterAdaptation,testLista)
                self.assertEqual(len(captured.records), 1)  # check that there is only one log message
                self.assertEqual(captured.records[0].getMessage(), "[TYPE ERROR] " + "Name of the parameter: " + self.Object.name + "  Type of the parameter: " + self.Object.type)

    def test_paramaterAdaptation_002(self):
        self.Object2 = ArchitectExcel.DataFromXlsArchitect("TEST2", "uint8", 1, -23456, "Infoblock", "PPAR_ChannelVariationData", "0xFFFF")
        self.Object = ArchitectExcel.DataFromXlsArchitect("TEST", "uint8", 1, "0x0", "Infoblock", "PPAR_ChannelVariationData", "0xFFFF")
        testLista = []
        testLista.append(self.Object)
        testLista.append(self.Object2)
        ArchitectExcel.parameterAdaptation(testLista)
        self.assertEqual(self.Object.type, "uint8")
        self.assertEqual(self.Object.value, "0x00")
        with open("LogWarning.txt") as f:
            contents = f.read()
        self.assertEqual(contents,"[ARCHITECT SECTION]\n[OUT OF RANGE]  Name: " + str(self.Object2.name) + ", Type: " + str(self.Object2.type) + ", Value: " + str(self.Object2.value) + "\n")

    def test_parameterVerificationSystem_001(self):
        self.Object2 = ArchitectExcel.DataFromXlsArchitect("TEST2", "uint8", 1, -23456, "Infoblock", "PPAR_ChannelVariationData", "0xFFFF")
        self.Object12345 = SystemExcel.DataFromXlsSystem2("TEST_System2", "SI8", 1, "upper", "0xFF","PPAR_InfoBlock", 100,200)
        self.Object1234 = SystemExcel.DataFromXlsSystem2("TEST_System1", "32", 1, "upper", "0x0000","PPAR_InfoBlock", 100,200)
        objTest = ArchitectExcel.DataFromXlsArchitect("TEST2.2", "32", 1, -23456, "Infoblock", "PPAR_ChannelVariationData", "0xFFFF")
        objTest2 = ArchitectExcel.DataFromXlsArchitect("TEST3.0", "sint8", 1, "0xFF", "Infoblock","PPAR_ChannelVariationData", "0xFFFF")
        testLista = [self.Object12345, self.Object1234]
        wx = GUI_Ppar.MyApp()
        SystemExcel.parameterVerificationSystem(testLista)
        with open("LogWarning.txt") as f:
            contents = f.read()
        self.assertEqual(contents,"[ARCHITECT SECTION]\n[OUT OF RANGE]  Name: " + str(self.Object2.name) + ", Type: " + str(self.Object2.type) + ", Value: " + str(self.Object2.value) + "\n\n[OUTPUT SECTION]\n" +"[TYPE ERROR] " + "Name of the parameter: " + objTest.name + "  Type of the parameter: " + objTest.type + "\n" + "\n[OUTPUT SECTION]\n" + "[OUT OF RANGE]  Name: " + str(objTest2.name) + ", Type: " + str(objTest2.type) + ", Value: " + str(objTest2.value) + "\n" +"\n[SYSTEM SECTION]\n[OUT OF RANGE]  Name: " + str(self.Object12345.name) + ", Type: " + str(self.Object12345.type) + ", Value: " + str(self.Object12345.value) + "\n[TYPE ERROR]    Name of the parameter: " + str(self.Object1234.name) + "   Type of the parameter: " + str(self.Object1234.type) + "\n")

    def test_parameterVerificationOutput_001(self):
        self.Object2 = ArchitectExcel.DataFromXlsArchitect("TEST2", "uint8", 1, -23456, "Infoblock","PPAR_ChannelVariationData", "0xFFFF")
        objTest = ArchitectExcel.DataFromXlsArchitect("TEST2.2", "32", 1, -23456, "Infoblock", "PPAR_ChannelVariationData", "0xFFFF")
        objTest1 = SystemExcel.DataFromXlsSystem2("TEST2.2", "uint16", 1, "upper", "0x0010","PPAR_InfoBlock", 100,200)
        listaArchitect = [objTest]
        listaSystem = [objTest1]
        wx = GUI_Ppar.MyApp()
        self.assertRaises(RuntimeError, SystemExcel.parameterVerificationOutput,listaArchitect,listaSystem)
        with open("LogWarning.txt") as f:
            contents = f.read()
            self.assertEqual(contents, "[ARCHITECT SECTION]\n[OUT OF RANGE]  Name: " + str(self.Object2.name) + ", Type: " + str(self.Object2.type) + ", Value: " + str(self.Object2.value) + "\n\n[OUTPUT SECTION]\n" + "[TYPE ERROR] " + "Name of the parameter: " + objTest.name + "  Type of the parameter: " + objTest.type + "\n")

    def test_parameterVerificationOutput_002(self):
        self.Object2 = ArchitectExcel.DataFromXlsArchitect("TEST2", "uint8", 1, -23456, "Infoblock","PPAR_ChannelVariationData", "0xFFFF")
        objTest = ArchitectExcel.DataFromXlsArchitect("TEST2.2", "32", 1, -23456, "Infoblock", "PPAR_ChannelVariationData", "0xFFFF")
        objTest1 = SystemExcel.DataFromXlsSystem2("TEST2.2", "uint16", 1, "upper", "0x0010","PPAR_InfoBlock", 100,200)
        objTest2 = ArchitectExcel.DataFromXlsArchitect("TEST3.0", "sint8", 1, "0xFF", "Infoblock","PPAR_ChannelVariationData", "0xFFFF")
        listaArchitect = [objTest2]
        listaSystem = [objTest1]
        wx = GUI_Ppar.MyApp()
        SystemExcel.parameterVerificationOutput(listaArchitect,listaSystem)
        with open("LogWarning.txt") as f:
            contents = f.read()
            self.assertEqual(contents, "[ARCHITECT SECTION]\n[OUT OF RANGE]  Name: " + str(self.Object2.name) + ", Type: " + str(self.Object2.type) + ", Value: " + str(self.Object2.value) + "\n\n[OUTPUT SECTION]\n" + "[TYPE ERROR] " + "Name of the parameter: " + objTest.name + "  Type of the parameter: " + objTest.type + "\n" + "\n[OUTPUT SECTION]\n" + "[OUT OF RANGE]  Name: " + str(objTest2.name) + ", Type: " + str(objTest2.type) + ", Value: " + str(objTest2.value) + "\n")

    def test_readxmlSystem_001(self):
        objectOem = OemSpecific.OemSpecific("a_GfaKey", "uint8", 16, "0xFF", "PPAR_SelfTest", "PPAR_OemSwBlock", None)
        objectOem2 = OemSpecific.OemSpecific("a_GfaKey", "uint8", 16, "0xFF", "PPAR_SelfTest", "PPAR_OemSwBlock", None)
        listOem = [objectOem,objectOem2]
        Object11 = DataFromXlsArchitect("res1", "UI8", 1, "0x0", "Infoblock", "PPAR_ChannelVariationData", None)
        global_.row = 4
        global_.nameColumn = 4
        global_.typeColumn = 17
        global_.countColumn = 18
        global_.valueColumn = 27
        global_.resColumn = 20
        global_.rowEnd = "PPAR_SECTION_END"
        global_.readRowAddress = 4
        global_.readColumnAddress = 10
        global_.rowA2L = 6
        global_.columnA2L = 9
        global_.defaultColumn = 23
        global_.excelReadArchitect = load_workbook(filename="UnitTest/ArchitectTEST.xlsx", data_only=True)
        global_.sheetArchitect = global_.excelReadArchitect[global_.architectSheet]
        wx = GUI_Ppar.MyApp()
        listt = ArchitectExcel.DataFromXlsArchitect.readXMLArchitect(Object11,listOem)
        for x in listt:
            if x.name == "ui8_InfoblockString":
                continue
            self.assertEqual(x.name, "ui16_LengthInfoblock")
            self.assertEqual(x.type, "UI16")
            self.assertEqual(x.count, 1)
            self.assertEqual(x.value, "0x80")
            self.assertEqual(x.L2Architect, "Infoblock")
            self.assertEqual(x.TreeLevel2, "CommonInfoBlock")
            break

    def test_writeInExcel_001(self):
        listF = []
        objectOem = OemSpecific.OemSpecific("a_GfaKey", "uint8", 16, "0xFF", "PPAR_SelfTest", "PPAR_OemSwBlock", None)
        objectOem2 = OemSpecific.OemSpecific("a_GfaKey", "uint8", 16, "0xFF", "PPAR_SelfTest", "PPAR_OemSwBlock", None)
        listOem = [objectOem, objectOem2]
        Object11 = DataFromXlsArchitect("res1", "UI8", 1, "0x0", "Infoblock", "PPAR_ChannelVariationData", None)
        global_.row = 4
        global_.nameColumn = 4
        global_.typeColumn = 17
        global_.rowWriteSheet2 = 2
        global_.countColumn = 18
        global_.valueColumn = 27
        global_.resColumn = 20
        global_.rowEnd = "PPAR_SECTION_END"
        global_.readRowAddress = 4
        global_.readColumnAddress = 10
        global_.rowA2L = 6
        global_.columnA2L = 9
        global_.defaultColumn = 23
        global_.nameColumnWrite = 5
        global_.typeColumnWrite = 9
        global_.countColumnWrite = 10
        global_.valueColumnWrite = 21
        global_.rowWrite = 6
        global_.writeL2Architect = 3
        global_.writeTreeLevel2 = 4
        global_.columnRes = 100
        global_.rowRes = 2
        global_.writeRowAddress = 5
        global_.writeColumnAddress = 148
        global_.columnWriteSheet2 = 2
        global_.rowWriteA2L = 6
        global_.columnWriteA2L = 152
        global_.lowLimitColumn = 20
        global_.maxLimitColumn = 22
        global_.excelWriteOutput = load_workbook(filename="UnitTest/xPAR_TEST.xlsx")
        global_.sheetOutput = global_.excelWriteOutput[global_.output1Sheet]
        global_.sheetOutput2 = global_.excelWriteOutput[global_.output2Sheet]
        global_.excelReadArchitect = load_workbook(filename="UnitTest/ArchitectTEST.xlsx",
                                                   data_only=True)
        global_.sheetArchitect = global_.excelReadArchitect[global_.architectSheet]
        # global_.sheetArchitect = "PPAR Definition"
        wx = GUI_Ppar.MyApp()
        listF = ArchitectExcel.DataFromXlsArchitect.readXMLArchitect(Object11,listOem)
        ArchitectExcel.DataFromXlsArchitect.writeInExcel(listF)

    #SystemExcel.py
    def test_repr_002(self):
        self.Object = SystemExcel.DataFromXlsSystem2("TEST", "UI8", 1, "upper", "0x0000","PPAR_InfoBlock", 100,200)
        print(self.Object)

    def test_readxmlSystem2_002(self):
        Object12 = SystemExcel.DataFromXlsSystem2("TEST", "UI8", 1, "upper", "0x0000","PPAR_InfoBlock", 100,200)
        global_.excelReadSystem = load_workbook(filename="UnitTest/xPAR-Definition_TEST.xlsx")
        global_.sheetSystem = global_.excelReadSystem[global_.systemSheet]
        global_.nameColumn2 = 2
        global_.typeColumn2 = 10
        global_.countColumn2 = 11
        global_.valueColumn2 = 20
        global_.row2 = 10
        global_.rowEnd2 = "si16_LoopbackPowerMax_dB"
        global_.projectSpecificName="ARS512VW13"
        global_.projectNameType="default ARS"
        global_.limitValueColumn = 19
        global_.listWithNotUsedElements="u_SubCompatID , u_ProjectCompatID, si16_LfSporadicNoiseMax"
        list = SystemExcel.DataFromXlsSystem2.readXMLSystem2(Object12)
        for x in list:
            self.assertEqual(x.name, "Attenuation STC at Corner")
            self.assertEqual(x.type, "UI16")
            self.assertEqual(x.count, 1)
            self.assertEqual(x.limit, "upper")
            self.assertEqual(x.value, "0x0001")
            break

    def test_writeInExcel2_001(self):
        # self.Object6 = SystemExcel.DataFromXlsSystem2("TEST", "UI8", 1, "upper", "0x0000")
        # list = []
        # list.append(self.Object6)
        objectOem = OemSpecific.OemSpecific("a_GfaKey", "uint8", 16, "0xFF", "PPAR_SelfTest", "PPAR_OemSwBlock", None)
        objectOem2 = OemSpecific.OemSpecific("a_GfaKey", "uint8", 16, "0xFF", "PPAR_SelfTest", "PPAR_OemSwBlock", None)
        listOem = [objectOem, objectOem2]
        Object11 = DataFromXlsArchitect("res1", "UI8", 1, "0x0", "Infoblock", "PPAR_ChannelVariationData", None)
        global_.row = 4
        global_.nameColumn = 4
        global_.typeColumn = 17
        global_.countColumn = 18
        global_.valueColumn = 27
        global_.resColumn = 20
        global_.rowEnd = "PPAR_SECTION_END"
        global_.readRowAddress = 4
        global_.readColumnAddress = 10
        global_.rowA2L = 6
        global_.columnA2L = 9
        global_.defaultColumn = 23
        global_.nameColumnWrite = 5
        global_.typeColumnWrite = 9
        global_.countColumnWrite = 10
        global_.valueColumnWrite = 21
        global_.rowWrite = 6
        global_.writeL2Architect = 3
        global_.writeTreeLevel2 = 4
        global_.columnRes = 100
        global_.rowRes = 2
        global_.writeRowAddress = 5
        global_.writeColumnAddress = 148
        global_.rowWriteSheet2 = 2
        global_.columnWriteSheet2 = 2
        global_.rowWriteA2L = 6
        global_.columnWriteA2L = 152
        global_.lowLimitColumn = 20
        global_.maxLimitColumn = 22
        global_.excelWriteOutput = load_workbook(filename="UnitTest/xPAR_TEST.xlsx")
        global_.sheetOutput = global_.excelWriteOutput[global_.output1Sheet]
        global_.sheetOutput2 = global_.excelWriteOutput[global_.output2Sheet]
        global_.excelReadArchitect = load_workbook(filename="UnitTest/ArchitectTEST.xlsx",
                                                   data_only=True)
        global_.sheetArchitect = global_.excelReadArchitect[global_.architectSheet]
        # global_.sheetArchitect = "PPAR Definition"
        wx = GUI_Ppar.MyApp()
        #listFArchitect = ArchitectExcel.DataFromXlsArchitect.readXMLArchitect(Object11,listOem)
        #ArchitectExcel.DataFromXlsArchitect.writeInExcel(listFArchitect)
        global_.excelReadSystem = load_workbook(filename="UnitTest/xPAR-Definition_TEST.xlsx")
        global_.sheetSystem = global_.excelReadSystem[global_.systemSheet]
        Object12 = SystemExcel.DataFromXlsSystem2("TEST", "UI8", 1, "upper", "0x0000", "PPAR_InfoBlock", 100, 200)
        global_.nameColumn2 = 2
        global_.typeColumn2 = 10
        global_.countColumn2 = 11
        global_.valueColumn2 = 20
        global_.row2 = 10
        global_.rowEnd2 = "si16_LoopbackPowerMax_dB"
        global_.projectSpecificName = "ARS512VW13"
        global_.projectNameType = "default ARS"
        global_.limitValueColumn = 19
        global_.listWithNotUsedElements = "u_SubCompatID , u_ProjectCompatID, si16_LfSporadicNoiseMax"
        wx = GUI_Ppar.MyApp()
        listF = SystemExcel.DataFromXlsSystem2.readXMLSystem2(Object12)
        SystemExcel.DataFromXlsSystem2.writeInExcel2(listF)

    def test_a2lFunction_001(self):
        global_.excelReadArchitect = load_workbook(filename="UnitTest/ArchitectTEST.xlsx",
                                                   data_only=True)
        global_.sheetArchitect = global_.excelReadArchitect[global_.architectSheet]
        global_.rowA2L = 6
        global_.nameColumn = 4
        global_.rowEnd = "PPAR_SECTION_END"
        global_.projectOemName = "PPAR_OemSwBlock_ARS512VW13"
        global_.rowOEMa2l = 3
        global_.columnOemA2L = 4
        global_.columnA2L = 9
        A2LColumn.a2lFunction()

    def test_a2lWrite_001(self):
        global_.excelReadArchitect = load_workbook(filename="UnitTest/ArchitectTEST.xlsx",
                                                   data_only=True)
        global_.sheetArchitect = global_.excelReadArchitect[global_.architectSheet]
        global_.rowA2L = 6
        global_.nameColumn = 4
        global_.rowEnd = "PPAR_SECTION_END"
        global_.projectOemName = "PPAR_OemSwBlock_ARS512VW13"
        global_.rowOEMa2l = 3
        global_.columnOemA2L = 4
        global_.columnA2L = 9
        listF = A2LColumn.a2lFunction()
        global_.rowWriteA2L = 6
        global_.columnWriteA2L = 152
        global_.excelWriteOutput = load_workbook(filename="UnitTest/xPAR_TEST.xlsx")
        global_.sheetOutput = global_.excelWriteOutput[global_.output1Sheet]
        global_.sheetOutput2 = global_.excelWriteOutput[global_.output2Sheet]
        A2LColumn.a2lWrite(listF)

    def test_reprOem_001(self):
        objectOem = OemSpecific.OemSpecific("a_GfaKey", "uint8", 16, "0xFF", "PPAR_SelfTest", "PPAR_OemSwBlock", None)
        print(objectOem)

    def test_readOemSpecific_001(self):
        objectOem = OemSpecific.OemSpecific("a_GfaKey", "uint8", 16, "0xFF", "PPAR_SelfTest", "PPAR_OemSwBlock", None)
        global_.excelReadArchitect = load_workbook(filename="UnitTest/ArchitectTEST.xlsx",
                                                   data_only=True)
        global_.sheetArchitect = global_.excelReadArchitect[global_.architectSheet]
        global_.projectOemName = "PPAR_OemSwBlock_ARSa342tsd=512VW13"
        global_.rowOem = 3
        global_.nameColumnOem = 2
        global_.countColumnOem = 10
        global_.valueColumnOem = 15
        global_.rowEndOem = "a_Reserved"
        global_.valueColumn = 27
        OemSpecific.OemSpecific.readOemSpecific(objectOem)

    def test_readOemSpecific_002(self):
        objectOem = OemSpecific.OemSpecific("a_GfaKey", "uint8", 16, "0xFF", "PPAR_SelfTest", "PPAR_OemSwBlock", None)
        global_.excelReadArchitect = load_workbook(filename="UnitTest/ArchitectTEST.xlsx",
                                                   data_only=True)
        global_.sheetArchitect = global_.excelReadArchitect[global_.architectSheet]
        global_.projectOemName = "PPAR_OemSwBlock_ARS512VW13"
        global_.rowOem = 3
        global_.nameColumnOem = 2
        global_.countColumnOem = 10
        global_.valueColumnOem = 15
        global_.rowEndOem = "a_Reserved"
        global_.valueColumn = 27
        OemSpecific.OemSpecific.readOemSpecific(objectOem)

if __name__ == '__main__':  # pragma: no cover
    unittest.main()  # pragma: no cover