from openpyxl import load_workbook
import global_


def a2lFunction():                              # copies the A2L names from Architect excel
    listDataFromA2L = []
    valid = 0

    listSheets = global_.excelReadArchitect.sheetnames

    for sheet in listSheets:
        if global_.projectOemName == sheet:
            sheetOem = global_.excelReadArchitect[sheet]
            valid = 1

    while global_.rowA2L:
        name = global_.sheetArchitect.cell(global_.rowA2L, global_.nameColumn).value
        if name == global_.rowEnd:
            global_.rowA2L = False
            break

        if name == "Placeholder for OemSwBlock" and valid == 1:
            while global_.rowOEMa2l:
                nameA2l = sheetOem.cell(global_.rowOEMa2l, global_.columnOemA2L).value
                global_.rowOEMa2l += 1
                listDataFromA2L.append(nameA2l)

                if nameA2l == "PPAR_OemSwBlock.a_Reserved" or nameA2l == "PPAR_OemSwBlock.aReserved":
                    global_.rowOEMa2l = False

        else:
            a2l = global_.sheetArchitect.cell(global_.rowA2L, global_.columnA2L).value
            if a2l:
                listDataFromA2L.append(a2l)
        global_.rowA2L = global_.rowA2L + 1
    return listDataFromA2L


def a2lWrite(list):                             # writes the A2L names to Output excel
    for iterator in list:
        global_.sheetOutput.cell(row=global_.rowWriteA2L, column=global_.columnWriteA2L, value=iterator)
        global_.rowWriteA2L = global_.rowWriteA2L + 1
    global_.excelWriteOutput.save(global_.pathOutput)
